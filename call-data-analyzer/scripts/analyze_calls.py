#\!/usr/bin/env python3
"""
Call Data Analyzer - Main analysis script
Analyzes CSV files with call data and generates Excel reports with sentiment analysis
"""

import pandas as pd
import json
import sys
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def analyze_sentiment(text):
    """Simple keyword-based sentiment analysis"""
    if not text or pd.isna(text):
        return 'Neutral'
    
    text_lower = str(text).lower()
    
    negative_words = [
        'frustrated', 'angry', 'upset', 'disappointed', 'unhappy',
        'problem', 'issue', 'failed', 'failed to', 'not resolved',
        'delay', 'wrong', 'missing', 'damaged', 'complaint',
        'escalate', 'escalation', 'demanded'
    ]
    
    positive_words = [
        'resolved', 'solved', 'helped', 'assistance',
        'satisfied', 'confirmed', 'provided', 'quickly'
    ]
    
    negative_count = sum(1 for word in negative_words if word in text_lower)
    positive_count = sum(1 for word in positive_words if word in text_lower)
    
    if negative_count > positive_count:
        return 'Negative'
    elif positive_count > negative_count:
        return 'Positive'
    else:
        return 'Neutral'

def parse_call_data(csv_files):
    """Parse CSV files and extract call data with sentiment analysis"""
    all_data = []
    
    for csv_file in csv_files:
        df = pd.read_csv(csv_file)
        
        # Detect format by checking columns
        if 'Summary' in df.columns:
            # Old format with JSON summary
            for idx, row in df.iterrows():
                try:
                    if pd.notna(row.get('Summary')):
                        summary = json.loads(row['Summary'])
                        csat_val = summary.get('Resolution Metrics', {}).get('CSAT Score')
                        try:
                            csat_val = float(csat_val) if csat_val else None
                        except:
                            csat_val = None
                        
                        summary_text = summary.get('Summary', '')
                        friction_text = summary.get('Performance Indicators', {}).get('Key Friction Points', '')
                        sentiment = analyze_sentiment(summary_text + ' ' + friction_text)
                        
                        all_data.append({
                            'Call_ID': row.get('Call ID'),
                            'Date': row.get('Date'),
                            'Duration_Seconds': row.get('Call Duration (Seconds)'),
                            'Issue_Category': summary.get('Resolution Metrics', {}).get('Issue Category'),
                            'Resolution_Status': summary.get('Resolution Metrics', {}).get('Resolution Status'),
                            'CSAT_Score': csat_val,
                            'Bot_Effectiveness': summary.get('Performance Indicators', {}).get('Bot Effectiveness'),
                            'Transfer_Reason': summary.get('Performance Indicators', {}).get('Transfer Reason'),
                            'Friction_Points': summary.get('Performance Indicators', {}).get('Key Friction Points'),
                            'Sentiment': sentiment,
                            'Summary': summary_text[:200]
                        })
                except Exception as e:
                    continue
        else:
            # New format with flat columns
            for idx, row in df.iterrows():
                try:
                    # Extract sentiment from existing sentiment column if present, or analyze from text
                    if 'sentiment' in df.columns and pd.notna(row.get('sentiment')):
                        sentiment = row['sentiment'].capitalize() if isinstance(row['sentiment'], str) else 'Neutral'
                    else:
                        summary_text = str(row.get('summary_text', '')) if pd.notna(row.get('summary_text')) else ''
                        friction_text = str(row.get('friction_points', '')) if pd.notna(row.get('friction_points')) else ''
                        sentiment = analyze_sentiment(summary_text + ' ' + friction_text)
                    
                    csat_val = row.get('csat')
                    try:
                        csat_val = float(csat_val) if pd.notna(csat_val) else None
                    except:
                        csat_val = None
                    
                    all_data.append({
                        'Call_ID': f"CALL_{idx+1}",  # Generate Call ID if not present
                        'Date': row.get('date', ''),
                        'Duration_Seconds': row.get('duration', ''),
                        'Issue_Category': row.get('issue_category', ''),
                        'Resolution_Status': row.get('resolution_status', ''),
                        'CSAT_Score': csat_val,
                        'Bot_Effectiveness': row.get('bot_effectiveness', ''),
                        'Transfer_Reason': row.get('transfer_reason', ''),
                        'Friction_Points': row.get('friction_points', ''),
                        'Sentiment': sentiment,
                        'Summary': str(row.get('summary_text', ''))[:200] if pd.notna(row.get('summary_text')) else ''
                    })
                except Exception as e:
                    continue
    
    return pd.DataFrame(all_data)

def create_excel_report(analysis_df, output_path):
    """Create multi-sheet Excel report with analysis"""
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Issue Type Analysis
    ws1 = wb.active
    ws1.title = "Issue Type Analysis"
    
    issue_analysis = []
    for issue in sorted(analysis_df['Issue_Category'].dropna().unique()):
        issue_calls = analysis_df[analysis_df['Issue_Category'] == issue]
        total = len(issue_calls)
        
        sentiment_dist = issue_calls['Sentiment'].value_counts()
        pos = sentiment_dist.get('Positive', 0)
        neu = sentiment_dist.get('Neutral', 0)
        neg = sentiment_dist.get('Negative', 0)
        
        csat_data = issue_calls[issue_calls['CSAT_Score'].notna()]['CSAT_Score']
        avg_csat = csat_data.mean() if len(csat_data) > 0 else None
        
        issue_analysis.append({
            'Issue_Type': issue,
            'Total_Calls': total,
            'Percentage': f"{total/len(analysis_df)*100:.2f}%",
            'Positive': pos,
            'Neutral': neu,
            'Negative': neg,
            'Neg_%': f"{neg/total*100:.1f}%" if total > 0 else "0%",
            'Avg_CSAT': f"{avg_csat:.2f}" if avg_csat else "N/A",
            'CSAT_Responses': len(csat_data)
        })
    
    issue_df = pd.DataFrame(issue_analysis).sort_values('Total_Calls', ascending=False)
    
    headers = list(issue_df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for row_num, row in enumerate(issue_df.values, 2):
        for col_num, value in enumerate(row, 1):
            cell = ws1.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
    
    ws1.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws1.column_dimensions[col].width = 14
    
    # Sheet 2: All Calls Data
    ws2 = wb.create_sheet("All Calls Data")
    export_cols = ['Call_ID', 'Date', 'Duration_Seconds', 'Issue_Category', 'Resolution_Status', 
                   'Sentiment', 'CSAT_Score', 'Bot_Effectiveness']
    export_df = analysis_df[export_cols].copy()
    
    for col_num, header in enumerate(export_cols, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for row_num, row in enumerate(export_df.values, 2):
        for col_num, value in enumerate(row, 1):
            cell = ws2.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            if col_num == 6:  # Sentiment
                if value == 'Negative':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif value == 'Positive':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    ws2.column_dimensions['A'].width = 30
    ws2.freeze_panes = 'A2'
    
    # Sheet 3: CSAT Analysis
    ws3 = wb.create_sheet("CSAT Analysis")
    csat_calls = analysis_df[analysis_df['CSAT_Score'].notna()][['Call_ID', 'Issue_Category', 'CSAT_Score', 'Sentiment']].sort_values('CSAT_Score')
    
    csat_headers = ['Call_ID', 'Issue_Category', 'CSAT_Score', 'Sentiment']
    for col_num, header in enumerate(csat_headers, 1):
        cell = ws3.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for row_num, row in enumerate(csat_calls.values, 2):
        for col_num, value in enumerate(row, 1):
            cell = ws3.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            if col_num == 3:
                if value <= 2:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif value >= 4:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 25
    
    # Sheet 4: Negative Sentiment
    ws4 = wb.create_sheet("Negative Sentiment")
    neg_calls = analysis_df[analysis_df['Sentiment'] == 'Negative'][['Call_ID', 'Date', 'Issue_Category', 'CSAT_Score', 'Resolution_Status', 'Friction_Points']].sort_values('Date', ascending=False)
    
    if len(neg_calls) > 0:
        neg_headers = ['Call_ID', 'Date', 'Issue_Category', 'CSAT_Score', 'Resolution_Status', 'Friction_Points']
        for col_num, header in enumerate(neg_headers, 1):
            cell = ws4.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.font = header_font
            cell.border = border
        
        for row_num, row in enumerate(neg_calls.values, 2):
            for col_num, value in enumerate(row, 1):
                cell = ws4.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        ws4.column_dimensions['A'].width = 30
        ws4.column_dimensions['F'].width = 50
    
    # Sheet 5: Summary Statistics
    ws5 = wb.create_sheet("Summary Statistics")
    
    total_calls = len(analysis_df)
    sentiment_counts = analysis_df['Sentiment'].value_counts()
    csat_data = analysis_df['CSAT_Score'].dropna()
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Calls', total_calls],
        ['Positive', sentiment_counts.get('Positive', 0)],
        ['Neutral', sentiment_counts.get('Neutral', 0)],
        ['Negative', sentiment_counts.get('Negative', 0)],
        ['Negative %', f"{sentiment_counts.get('Negative', 0)/total_calls*100:.1f}%"],
        ['Avg CSAT', f"{csat_data.mean():.2f}" if len(csat_data) > 0 else 'N/A'],
        ['Satisfied (4-5)', len(csat_data[csat_data >= 4]) if len(csat_data) > 0 else 0],
        ['Dissatisfied (1-2)', len(csat_data[csat_data < 3]) if len(csat_data) > 0 else 0],
    ]
    
    for row_num, row in enumerate(summary_data, 1):
        for col_num, value in enumerate(row, 1):
            cell = ws5.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if row_num == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 15
    
    wb.save(output_path)

def main():
    csv_files = []
    output_path = "call_analysis.xlsx"
    
    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == '--output' and i + 1 < len(args):
            output_path = args[i + 1]
            i += 2
        elif not args[i].startswith('--'):
            csv_files.append(args[i])
            i += 1
        else:
            i += 1
    
    if not csv_files:
        print("Usage: python analyze_calls.py <csv_file1> [csv_file2] ... [--output output.xlsx]")
        sys.exit(1)
    
    print(f"Analyzing {len(csv_files)} CSV file(s)...")
    df = parse_call_data(csv_files)
    print(f"Parsed {len(df)} calls")
    print(f"Creating Excel report...")
    create_excel_report(df, output_path)
    print(f"Analysis complete\\! Report saved to {output_path}")

if __name__ == "__main__":
    main()
